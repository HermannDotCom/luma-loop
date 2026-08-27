"""Build a transparent 12-month Android rewarded-ad revenue scenario model for Luma Loop."""

from __future__ import annotations

import json
from pathlib import Path

import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = PROJECT_ROOT / "financials"
WORKBOOK_PATH = OUTPUT_DIR / "luma_loop_ad_forecast.xlsx"
CHART_PATH = OUTPUT_DIR / "luma_loop_ad_forecast.png"

DAYS_PER_MONTH = 30
SCENARIOS = {
    "Prudent": {
        "installs_m1": 100,
        "installs_m12": 800,
        "monthly_returning_players": 0.05,
        "dau_mau_ratio": 0.08,
        "rewarded_opportunities_per_dau": 0.30,
        "fill_rate": 0.80,
        "rewarded_ecpm_usd": 3.00,
    },
    "Central": {
        "installs_m1": 500,
        "installs_m12": 5_000,
        "monthly_returning_players": 0.10,
        "dau_mau_ratio": 0.10,
        "rewarded_opportunities_per_dau": 0.60,
        "fill_rate": 0.90,
        "rewarded_ecpm_usd": 5.10,
    },
    "Haut": {
        "installs_m1": 2_000,
        "installs_m12": 25_000,
        "monthly_returning_players": 0.20,
        "dau_mau_ratio": 0.15,
        "rewarded_opportunities_per_dau": 1.00,
        "fill_rate": 0.95,
        "rewarded_ecpm_usd": 8.00,
    },
}

TITLE_FILL = PatternFill("solid", fgColor="135B44")
SECTION_FILL = PatternFill("solid", fgColor="CFE9E0")
GRAY_FILL = PatternFill("solid", fgColor="E7E5E4")
WHITE_FONT = Font(color="FFFFFF", bold=True)
INPUT_FONT = Font(color="0000FF")
FORMULA_FONT = Font(color="000000")
CROSS_SHEET_FORMULA_FONT = Font(color="008000")
THIN_GRAY = Side(style="thin", color="B7B7B7")
TOP_BORDER = Border(top=THIN_GRAY)
DOUBLE_BOTTOM = Border(bottom=Side(style="double", color="000000"))


def scenario_projection(inputs: dict[str, float]) -> list[dict[str, float]]:
    """Return 12 monthly projections using the same equations written to the workbook."""
    rows: list[dict[str, float]] = []
    prior_mau = 0.0
    for month in range(1, 13):
        new_installs = inputs["installs_m1"] + (
            (inputs["installs_m12"] - inputs["installs_m1"]) * (month - 1) / 11
        )
        mau = new_installs + prior_mau * inputs["monthly_returning_players"]
        dau = mau * inputs["dau_mau_ratio"]
        impressions = (
            dau
            * DAYS_PER_MONTH
            * inputs["rewarded_opportunities_per_dau"]
            * inputs["fill_rate"]
        )
        revenue = impressions * inputs["rewarded_ecpm_usd"] / 1_000
        rows.append(
            {
                "month": month,
                "new_installs": new_installs,
                "mau": mau,
                "dau": dau,
                "filled_rewarded_impressions": impressions,
                "revenue_usd": revenue,
            }
        )
        prior_mau = mau
    return rows


def setup_sheet(ws, title: str, subtitle: str, unit: str) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws["C3"] = title
    ws["C3"].fill = TITLE_FILL
    ws["C3"].font = Font(color="FFFFFF", bold=True, size=16)
    ws.merge_cells("C3:H3")
    ws["C5"] = subtitle
    ws["C5"].font = Font(bold=True, size=11)
    ws.merge_cells("C5:H5")
    ws["C6"] = unit
    ws["C6"].font = Font(italic=True, color="666666")
    ws.merge_cells("C6:H6")
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = "B2:H60"
    ws.oddFooter.center.text = f"{ws.title} — Page &P"


def style_header(cells) -> None:
    for cell in cells:
        cell.fill = SECTION_FILL
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = TOP_BORDER


def comment_for(cell, text: str) -> None:
    cell.comment = Comment(text, "Manus AI")


def add_assumptions_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Hypothèses"
    setup_sheet(
        ws,
        "Luma Loop — Prévisionnel IAA Android",
        "Scénarios de revenus publicitaires récompensés — 12 mois",
        "USD, avant impôts, coûts d’acquisition, salaires et autres frais de Flash Digital SAS",
    )
    headers = ["Indicateur", "Prudent", "Central", "Haut", "Méthode / source"]
    for col, value in enumerate(headers, start=3):
        ws.cell(row=8, column=col, value=value)
    style_header(ws[8][2:8])
    rows = [
        ("Nouvelles installations — mois 1", "installs_m1", "#,#0", "Hypothèse de distribution, sans campagne payante incluse."),
        ("Nouvelles installations — mois 12", "installs_m12", "#,#0", "Hypothèse de distribution, sans campagne payante incluse."),
        ("Part des MAU précédents encore actifs", "monthly_returning_players", "0.0%", "Hypothèse de rétention mensuelle simplifiée; non issue de données Luma Loop."),
        ("DAU / MAU", "dau_mau_ratio", "0.0%", "Hypothèse de fréquence quotidienne; à remplacer par les données de production."),
        ("Impressions récompensées par DAU / jour", "rewarded_opportunities_per_dau", "0.00", "Hypothèse de placement volontaire; aucune publicité n’est active dans la version 1.0.0."),
        ("Taux de remplissage", "fill_rate", "0.0%", "Hypothèse de remplissage; dépend du pays, de la médiation et de la demande."),
        ("eCPM rewarded", "rewarded_ecpm_usd", "$#,##0.00;($#,##0.00);-", "Scénario central : benchmark Android Europe de 5,10 USD; bornes prudente/haute : hypothèses de sensibilité."),
        ("Jours par mois", None, "0", "Convention de modélisation uniforme sur 12 mois."),
    ]
    for row, (label, key, number_format, note) in enumerate(rows, start=9):
        ws.cell(row=row, column=3, value=label)
        for offset, (_, values) in enumerate(SCENARIOS.items(), start=4):
            cell = ws.cell(row=row, column=offset)
            if key is None:
                cell.value = DAYS_PER_MONTH
                comment_for(cell, "Convention de modélisation : 30 jours par mois, 27 août 2026.")
            else:
                cell.value = values[key]
                comment_for(cell, f"Source : {note} Date : 27 août 2026.")
            cell.font = INPUT_FONT
            cell.number_format = number_format
            cell.alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=7, value=note)
        ws.cell(row=row, column=7).alignment = Alignment(wrap_text=True, vertical="top")
    ws["C19"] = "Repère de marché"
    ws["C19"].fill = SECTION_FILL
    ws["C19"].font = Font(bold=True)
    ws.merge_cells("C19:H19")
    ws["C20"] = "Rétention mobile de référence"
    ws["D20"] = "Europe, médiane : D1 21,4 %, D7 4,31 %, D28 1,21 %"
    ws["D20"].font = INPUT_FONT
    ws.merge_cells("D20:H20")
    comment_for(ws["D20"], "Source : GameAnalytics, 2025 Mobile Gaming Benchmarks, données 2024, région Europe, consulté le 27 août 2026.")
    ws["C21"] = "eCPM Android Europe"
    ws["D21"] = "5,10 USD pour vidéo récompensée selon le benchmark publié"
    ws["D21"].font = INPUT_FONT
    ws.merge_cells("D21:H21")
    comment_for(ws["D21"], "Source : Mistplay, Mobile ads eCPM: Basics and latest data, 13 mars 2026, tableau Android Europe.")
    ws["C23"] = "Formule de revenus"
    ws["C23"].fill = SECTION_FILL
    ws["C23"].font = Font(bold=True)
    ws.merge_cells("C23:H23")
    ws["C24"] = "Revenu mensuel"
    ws["D24"] = "DAU × 30 jours × opportunités rewarded/DAU/jour × fill rate × eCPM / 1 000"
    ws.merge_cells("D24:H24")
    ws["C25"] = "Périmètre"
    ws["D25"] = "Modèle Android, publicité récompensée seulement, avant impôts et sans dépenses d’acquisition utilisateurs."
    ws.merge_cells("D25:H25")
    for col in range(3, 9):
        ws.column_dimensions[get_column_letter(col)].width = 24 if col != 7 else 54


def add_forecast_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Prévision")
    setup_sheet(
        ws,
        "Luma Loop — Prévision de revenus IAA",
        "Calcul mensuel par scénario — hypothèses liées à l’onglet Hypothèses",
        "USD, revenus publicitaires estimés",
    )
    scenario_columns = {"Prudent": "D", "Central": "D", "Haut": "D"}
    start_rows = {"Prudent": 9, "Central": 25, "Haut": 41}
    for scenario, start_row in start_rows.items():
        ws.cell(row=start_row - 1, column=3, value=scenario)
        ws.cell(row=start_row - 1, column=3).fill = SECTION_FILL
        ws.cell(row=start_row - 1, column=3).font = Font(bold=True)
        ws.merge_cells(start_row=start_row - 1, start_column=3, end_row=start_row - 1, end_column=8)
        headers = ["Mois", "Nouvelles installations", "MAU estimés", "DAU estimés", "Impressions rewarded", f"Revenus IAA — {scenario}"]
        for col, header in enumerate(headers, start=3):
            ws.cell(row=start_row, column=col, value=header)
        style_header(ws[start_row][2:8])
        assumption_col = {"Prudent": "D", "Central": "E", "Haut": "F"}[scenario]
        for month in range(1, 13):
            row = start_row + month
            ws.cell(row=row, column=3, value=month)
            ws.cell(row=row, column=4, value=f"=Hypothèses!{assumption_col}9+(Hypothèses!{assumption_col}10-Hypothèses!{assumption_col}9)*($C{row}-1)/11")
            if month == 1:
                ws.cell(row=row, column=5, value=f"=D{row}")
            else:
                ws.cell(row=row, column=5, value=f"=D{row}+E{row-1}*Hypothèses!{assumption_col}11")
            ws.cell(row=row, column=6, value=f"=E{row}*Hypothèses!{assumption_col}12")
            ws.cell(row=row, column=7, value=f"=F{row}*Hypothèses!{assumption_col}13*Hypothèses!{assumption_col}14*Hypothèses!{assumption_col}16")
            ws.cell(row=row, column=8, value=f"=G{row}*Hypothèses!{assumption_col}15/1000")
            for col in range(4, 9):
                cell = ws.cell(row=row, column=col)
                cell.font = CROSS_SHEET_FORMULA_FONT if col in (4, 5, 6, 7, 8) else FORMULA_FONT
                cell.alignment = Alignment(horizontal="right")
            for col in (4, 5, 6, 7):
                ws.cell(row=row, column=col).number_format = "#,##0"
            ws.cell(row=row, column=8).number_format = "$#,##0.00;($#,##0.00);-"
        total_row = start_row + 13
        ws.cell(row=total_row, column=3, value="Total / moyenne")
        ws.cell(row=total_row, column=4, value=f"=SUM(D{start_row + 1}:D{start_row + 12})")
        ws.cell(row=total_row, column=5, value=f"=AVERAGE(E{start_row + 1}:E{start_row + 12})")
        ws.cell(row=total_row, column=6, value=f"=AVERAGE(F{start_row + 1}:F{start_row + 12})")
        ws.cell(row=total_row, column=7, value=f"=SUM(G{start_row + 1}:G{start_row + 12})")
        ws.cell(row=total_row, column=8, value=f"=SUM(H{start_row + 1}:H{start_row + 12})")
        for col in range(3, 9):
            cell = ws.cell(row=total_row, column=col)
            cell.font = Font(bold=True, color="008000" if col > 3 else "000000")
            cell.border = Border(top=THIN_GRAY, bottom=Side(style="double", color="000000"))
            if col > 3:
                cell.alignment = Alignment(horizontal="right")
        for col in (4, 5, 6, 7):
            ws.cell(row=total_row, column=col).number_format = "#,##0"
        ws.cell(row=total_row, column=8).number_format = "$#,##0.00;($#,##0.00);-"
    for col in range(3, 9):
        ws.column_dimensions[get_column_letter(col)].width = 22
    chart = LineChart()
    chart.title = "Revenus IAA mensuels estimés"
    chart.y_axis.title = "USD"
    chart.x_axis.title = "Mois"
    chart.height = 8
    chart.width = 16
    for scenario, start_row in start_rows.items():
        data = Reference(ws, min_col=8, min_row=start_row, max_row=start_row + 12)
        categories = Reference(ws, min_col=3, min_row=start_row + 1, max_row=start_row + 12)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
    ws.add_chart(chart, "J9")


def add_summary_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Synthèse")
    setup_sheet(
        ws,
        "Luma Loop — Synthèse du prévisionnel IAA",
        "Vue de décision : publicité récompensée Android uniquement",
        "USD, avant impôts et dépenses de Flash Digital SAS",
    )
    headers = ["Indicateur", "Prudent", "Central", "Haut"]
    for col, header in enumerate(headers, start=3):
        ws.cell(row=8, column=col, value=header)
    style_header(ws[8][2:6])
    values = [
        ("Installations cumulées à 12 mois", "=Prévision!D22", "=Prévision!D38", "=Prévision!D54", "#,##0"),
        ("DAU moyen à 12 mois", "=Prévision!F22", "=Prévision!F38", "=Prévision!F54", "#,##0"),
        ("Revenu IAA total, 12 mois", "=Prévision!H22", "=Prévision!H38", "=Prévision!H54", "$#,##0.00;($#,##0.00);-"),
        ("Revenu IAA du mois 12", "=Prévision!H21", "=Prévision!H37", "=Prévision!H53", "$#,##0.00;($#,##0.00);-"),
        ("DAU requis pour 1 000 USD / mois", "=1000*1000/(Hypothèses!D13*Hypothèses!D14*Hypothèses!D15*Hypothèses!D16)", "=1000*1000/(Hypothèses!E13*Hypothèses!E14*Hypothèses!E15*Hypothèses!E16)", "=1000*1000/(Hypothèses!F13*Hypothèses!F14*Hypothèses!F15*Hypothèses!F16)", "#,##0"),
        ("DAU requis pour 10 000 USD / mois", "=D13*10", "=E13*10", "=F13*10", "#,##0"),
    ]
    for row, (label, prudent, central, high, number_format) in enumerate(values, start=9):
        ws.cell(row=row, column=3, value=label)
        for col, formula in enumerate([prudent, central, high], start=4):
            cell = ws.cell(row=row, column=col, value=formula)
            cell.font = CROSS_SHEET_FORMULA_FONT
            cell.number_format = number_format
            cell.alignment = Alignment(horizontal="right")
        if row in (11, 12):
            for col in range(3, 7):
                ws.cell(row=row, column=col).font = Font(bold=True, color="008000" if col > 3 else "000000")
                ws.cell(row=row, column=col).border = TOP_BORDER
    ws["C17"] = "Lecture de décision"
    ws["C17"].fill = SECTION_FILL
    ws["C17"].font = Font(bold=True)
    ws.merge_cells("C17:F17")
    ws["C18"] = "La publicité récompensée est une option de monétisation, pas une garantie de revenus. Dans ce modèle, le principal levier est le DAU durable : les revenus ne deviennent significatifs qu’à une audience régulière de plusieurs milliers de DAU."
    ws.merge_cells("C18:F19")
    ws["C18"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["C21"] = "Recommandation produit"
    ws["C21"].fill = SECTION_FILL
    ws["C21"].font = Font(bold=True)
    ws.merge_cells("C21:F21")
    ws["C22"] = "Publier 1.0.0 sans publicité, mesurer acquisition et rétention, puis tester dans une mise à jour une seule publicité récompensée volontaire à un moment de pause naturel. Éviter bannières et interstitiels dans la boucle de jeu."
    ws.merge_cells("C22:F23")
    ws["C22"].alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(3, 7):
        ws.column_dimensions[get_column_letter(col)].width = 28


def build_chart(projections: dict[str, list[dict[str, float]]]) -> None:
    plt.style.use("seaborn-v0_8-whitegrid")
    fig, (ax_full, ax_zoom) = plt.subplots(1, 2, figsize=(13, 5.5), dpi=180)
    colors = {"Prudent": "#7C8B94", "Central": "#43F3C5", "Haut": "#8B5CF6"}
    for scenario, rows in projections.items():
        months = [row["month"] for row in rows]
        revenue = [row["revenue_usd"] for row in rows]
        ax_full.plot(
            months,
            revenue,
            marker="o",
            linewidth=2.5,
            label=scenario,
            color=colors[scenario],
        )
        if scenario != "Haut":
            ax_zoom.plot(
                months,
                revenue,
                marker="o",
                linewidth=2.5,
                label=scenario,
                color=colors[scenario],
            )
    ax_full.set_title("Les trois scénarios (échelle logarithmique)", fontweight="bold")
    ax_full.set_xlabel("Mois après lancement")
    ax_full.set_ylabel("Revenu IAA estimé (USD)")
    ax_full.set_yscale("log")
    ax_full.set_xticks(range(1, 13))
    ax_full.legend(title="Scénario")
    ax_zoom.set_title("Prudent et central (échelle linéaire)", fontweight="bold")
    ax_zoom.set_xlabel("Mois après lancement")
    ax_zoom.set_ylabel("Revenu IAA estimé (USD)")
    ax_zoom.set_xticks(range(1, 13))
    ax_zoom.legend(title="Scénario")
    fig.suptitle("Luma Loop — Revenus publicitaires Android estimés", fontweight="bold", fontsize=14)
    fig.tight_layout(rect=(0, 0, 1, 0.94))
    fig.savefig(CHART_PATH, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def main() -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)
    projections = {name: scenario_projection(inputs) for name, inputs in SCENARIOS.items()}
    wb = Workbook()
    add_assumptions_sheet(wb)
    add_forecast_sheet(wb)
    add_summary_sheet(wb)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(WORKBOOK_PATH)
    build_chart(projections)
    summary = {
        scenario: {
            "installs_12_months": sum(row["new_installs"] for row in rows),
            "average_dau": sum(row["dau"] for row in rows) / len(rows),
            "revenue_12_months_usd": sum(row["revenue_usd"] for row in rows),
            "revenue_month_12_usd": rows[-1]["revenue_usd"],
            "dau_for_1000_usd_month": 1_000_000
            / (
                SCENARIOS[scenario]["rewarded_opportunities_per_dau"]
                * SCENARIOS[scenario]["fill_rate"]
                * SCENARIOS[scenario]["rewarded_ecpm_usd"]
                * DAYS_PER_MONTH
            ),
        }
        for scenario, rows in projections.items()
    }
    print(json.dumps({"workbook": str(WORKBOOK_PATH), "chart": str(CHART_PATH), "summary": summary}, indent=2))


if __name__ == "__main__":
    main()
